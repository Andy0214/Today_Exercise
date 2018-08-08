#! /usr/bin/env python
# -*- coding: utf-8 -*-
# author = "Andy"
# Date: 2018/8/7

# -*- coding:utf-8  -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import requests
import re
import urllib


def Get_Excel_data():
    file = './ip.xlsx'
    inwb = load_workbook(file)  # 读文件
    sheetnames = inwb.get_sheet_names()  # 获取文件中的所有的sheet表，通过名字的方式
    print(sheetnames)
    ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet表的内容
    rows = ws.max_row  # 获取读取的execl表的文件的行数
    print(rows)

    outwb = Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(title="cool")  # 在将写的文件中创建sheet表

    for i in range(rows):
        datavalues = ws.cell(row=i + 1, column=1).value
        print('已扫描 ' + str(i) + ' 个')
        URL = 'http://www.ip138.com/ips138.asp?ip='
        try:
            content = urllib.urlopen(URL + str(datavalues)).read()
            strdata = re.findall(r'<li>(.*?)</li>', content)
            if len(strdata):
                outws.cell(row=i + 1, column=1).value = datavalues
                outws.cell(row=i + 1, column=2).value = strdata[0].decode('gbk').encode('utf-8')
            else:
                outws.cell(row=i + 1, column=1).value = datavalues
        except requests.RequestException as e:
            outws.cell(row=i + 1, column=1).value = datavalues
            print(e)
    outwb.save('new_ip.xlsx')


Get_Excel_data()
